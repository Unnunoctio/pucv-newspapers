import os
import re
from datetime import datetime
from typing import List

from openpyxl import Workbook, load_workbook

from core.models import Article
from utils.logger import Logger


class ExcelExporter:
    @classmethod
    def _clean_text_to_excel(cls, text: str) -> str:
        if not text:
            return ""

        # Eliminar caracteres de control ASCII
        text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]", "", text)
        # Eliminar algunos caracteres invisibles Unicode problemáticos
        text = re.sub(r"[\u200b\u200c\u200d\u2028\u2029]", "", text)
        return text

    @staticmethod
    def export(articles: List[Article], file_name: str, folder_path: str) -> None:
        """Exporta los artículos a un archivo Excel"""
        if not articles:
            Logger.info("FILE", "Don't found any article to export")
            return
        
        file_path = os.path.join(folder_path, file_name)

        
        try:
            wb = load_workbook(file_path)
        except Exception:
            wb = Workbook()
            # Remove a default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        sheet_name = articles[0].newspaper.value

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["Newspaper", "URL", "Titulo", "Autor/Autores", "Fecha", "Tag", "Bajada", "Cuerpo", "Cuerpo HTML"])
        
        articles_sorted = sorted(articles, key=lambda p: p.date or datetime.min)

        for article in articles_sorted:
            ws.append(
                [
                    article.newspaper.value,
                    article.url,
                    ExcelExporter._clean_text_to_excel(article.title) or "",
                    ExcelExporter._clean_text_to_excel(article.author) or "",
                    article.date.strftime("%d-%m-%Y") if article.date else "",
                    ExcelExporter._clean_text_to_excel(article.tag) or "",
                    ExcelExporter._clean_text_to_excel(article.drophead) or "",
                    ExcelExporter._clean_text_to_excel(article.body) or "",
                    ExcelExporter._clean_text_to_excel(article.body_html) or "",
                ]
            )

        try:
            wb.save(file_path)
            Logger.info("FILE", f"Articles exported satisfactorily to {file_name}")
        except Exception as e:
            Logger.error("FILE", f"Error exporting articles to Excel: {e}")
